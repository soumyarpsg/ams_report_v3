"""Download helpers — CSV / Excel / PDF for the Loyalty KPI page tables.

Public API:
    render_download_row(df, base_name, key, sheet_name="Data")

Renders three compact download buttons styled by ``components/styles.py``.
File names embed the current tier filter and a yyyymmdd_hhmm timestamp so
re-downloads don't overwrite each other.

Dependencies:
    - openpyxl (already in requirements)
    - reportlab (added in this update — used only when the PDF button is
      pressed; import is deferred so the dashboard still runs if reportlab
      is missing — the PDF button just shows a one-time installation hint).
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Optional

import pandas as pd
import streamlit as st

from util_filters import get_selected_tier


# ---------------------------------------------------------------------------
# File-name helpers
# ---------------------------------------------------------------------------
def _slugify(text: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return s or "data"


def _build_filename(base_name: str, ext: str) -> str:
    tier = get_selected_tier() or "AllTiers"
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return f"{_slugify(base_name)}_{_slugify(tier)}_{ts}.{ext}"


# ---------------------------------------------------------------------------
# Format converters
# ---------------------------------------------------------------------------
def _csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def _xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    buf = io.BytesIO()
    # openpyxl handles sheet name length cap of 31 chars
    safe_sheet = (sheet_name or "Data")[:31]
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=safe_sheet, index=False)
        ws = writer.sheets[safe_sheet]
        # Auto-fit column widths (rough heuristic)
        for col_idx, col in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col)),
                df[col].astype(str).map(len).max() if not df.empty else 0,
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max(10, max_len + 2), 40
            )
        # Header row styling
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill("solid", fgColor="1F1F2E")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.freeze_panes = "A2"
    return buf.getvalue()


def _pdf_bytes(df: pd.DataFrame, title: str) -> Optional[bytes]:
    """Render the dataframe as a landscape A4 PDF table. Returns None when
    reportlab is missing (caller surfaces a friendly message)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
    except ImportError:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()
    h_style = ParagraphStyle(
        "h", parent=styles["Heading2"],
        fontName="Helvetica-Bold", fontSize=13, textColor=colors.HexColor("#1F1F2E"),
        spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#6B6B7B"),
        spaceAfter=6,
    )

    tier = get_selected_tier() or "All Tiers"
    elements = [
        Paragraph(f"Spencer's MSR — {title}", h_style),
        Paragraph(
            f"Tier filter: <b>{tier}</b> &nbsp;·&nbsp; Generated: "
            f"{datetime.now().strftime('%d-%b-%Y %H:%M')}",
            sub_style,
        ),
        Spacer(1, 2 * mm),
    ]

    # Cap rows for PDF readability — first 400 rows + a footnote.
    MAX_ROWS = 400
    show_df = df.head(MAX_ROWS).copy()
    truncated = len(df) > MAX_ROWS

    # Build table data: header row + body rows. Stringify everything so
    # reportlab doesn't choke on mixed types.
    header = [str(c) for c in show_df.columns]
    body = show_df.astype(object).where(pd.notna(show_df), "").astype(str).values.tolist()
    data = [header] + body

    # Sensible column widths: even split, capped per total page width
    total_w = landscape(A4)[0] - 20 * mm
    n_cols = max(1, len(header))
    col_w = total_w / n_cols

    table = Table(data, repeatRows=1, colWidths=[col_w] * n_cols)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F1F2E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8D8E0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
            [colors.white, colors.HexColor("#FAFAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)

    if truncated:
        elements.append(Spacer(1, 4 * mm))
        elements.append(Paragraph(
            f"<i>Showing first {MAX_ROWS:,} of {len(df):,} rows. "
            "Download the Excel or CSV file for the complete dataset.</i>",
            sub_style,
        ))

    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
def render_download_row(
    df: pd.DataFrame,
    base_name: str,
    key: str,
    sheet_name: str = "Data",
    pdf_title: Optional[str] = None,
) -> None:
    """Render a compact 3-button row: CSV · Excel · PDF.

    Args:
        df: The data to export (already filtered/aggregated for display).
        base_name: Human-readable label used in the filename (e.g.
            "Enrolment_Trend_By_Month").
        key: Unique key prefix per call-site (prevents Streamlit widget
            key collisions when many tables appear on one page).
        sheet_name: Excel sheet name (truncated to 31 chars).
        pdf_title: Heading printed at the top of the PDF. Defaults to
            ``base_name`` with underscores replaced.
    """
    if df is None or df.empty:
        st.caption("_No data to download for this section under the current filters._")
        return

    if pdf_title is None:
        pdf_title = base_name.replace("_", " ").strip()

    st.markdown('<div class="download-row">', unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 1, 1])

    with c1:
        st.download_button(
            "⬇ CSV",
            data=_csv_bytes(df),
            file_name=_build_filename(base_name, "csv"),
            mime="text/csv",
            key=f"dl_csv_{key}",
            use_container_width=True,
        )
    with c2:
        st.download_button(
            "⬇ Excel",
            data=_xlsx_bytes(df, sheet_name=sheet_name),
            file_name=_build_filename(base_name, "xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_xlsx_{key}",
            use_container_width=True,
        )
    with c3:
        pdf_data = _pdf_bytes(df, pdf_title)
        if pdf_data is None:
            # reportlab not installed — show a disabled-looking button.
            st.button(
                "⬇ PDF (install reportlab)",
                key=f"dl_pdf_disabled_{key}",
                use_container_width=True,
                disabled=True,
                help=(
                    "PDF export requires the `reportlab` package. Install with: "
                    "`pip install reportlab>=4.0.0` and restart Streamlit."
                ),
            )
        else:
            st.download_button(
                "⬇ PDF",
                data=pdf_data,
                file_name=_build_filename(base_name, "pdf"),
                mime="application/pdf",
                key=f"dl_pdf_{key}",
                use_container_width=True,
            )

    st.markdown('</div>', unsafe_allow_html=True)
